"""
Экспорт транзакций в CSV и Excel.
Показывает работу с Pandas: создание DataFrame, форматирование, запись файлов.
"""
from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd

from database.queries import get_transactions, TransactionRow


def transactions_to_dataframe(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    type_filter: Optional[str] = None,
) -> pd.DataFrame:
    """Загружает транзакции в Pandas DataFrame."""
    rows: list[TransactionRow] = get_transactions(
        date_from=date_from,
        date_to=date_to,
        type_filter=type_filter,
        limit=10_000,   # для экспорта снимаем ограничение
    )

    if not rows:
        return pd.DataFrame(columns=["Дата", "Тип", "Сумма", "Категория", "Заметка"])

    data = [
        {
            "Дата": r.date.strftime("%d.%m.%Y"),
            "Тип": "Доход" if r.type == "income" else "Расход",
            "Сумма": r.amount,
            "Категория": r.category_name,
            "Заметка": r.note,
        }
        for r in rows
    ]

    df = pd.DataFrame(data)
    return df


def export_to_csv(path: str | Path, **filters) -> int:
    """Сохраняет транзакции в CSV. Возвращает кол-во строк."""
    df = transactions_to_dataframe(**filters)
    df.to_csv(path, index=False, encoding="utf-8-sig")   # utf-8-sig для Excel
    return len(df)


def export_to_excel(path: str | Path, **filters) -> int:
    """
    Сохраняет транзакции в .xlsx с форматированием:
    - заголовки жирным
    - числа с двумя знаками
    - автоширина колонок
    """
    df = transactions_to_dataframe(**filters)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Транзакции")

        ws = writer.sheets["Транзакции"]

        # Жирные заголовки
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="3D405B")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Автоширина колонок
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # Формат числа для колонки "Сумма" (3-я колонка)
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = "#,##0.00"

    return len(df)
